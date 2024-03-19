'''
Created on 18 Jan 2024

@author: matze
'''
#install python-openpyxl (arch native) or pylightxl which handles Softmaker better
# pip install xlrd if xls files are necessary - else let them convert.. 
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
from datetime import timedelta
from time import sleep
from enum import Enum
from collections import Counter
import os,sys
from itertools import tee
import gzip,time
import logging
from logging.handlers import RotatingFileHandler
import subprocess
from subprocess import Popen

class OSTools():
    #singleton, class methods only

    @classmethod
    def getLocalPath(cls,fileInstance):
        return os.path.dirname(os.path.realpath(fileInstance))
    
    @classmethod
    def fileExists(cls, path):
        return os.path.isfile(path)
    
    @classmethod
    def setMainWorkDir(cls,dirpath):
        os.chdir(dirpath)  #changes the "active directory" 
    
    @classmethod    
    def getActiveDirectory(cls):
        return os.getcwd()
    
    #@classmethod
    #def username(cls):
    #    return pwd.getpwuid(os.getuid()).pw_name 
    
    @classmethod
    def joinPathes(cls,*pathes):
        res=pathes[0]
        for _,tail in cls.__pairwise(pathes):
        #for a, b in tee(pathes):
            res = os.path.join(res, tail)
        return res
    
    @classmethod
    def getHomeDirectory(cls):
        return os.path.expanduser("~")
    
    @classmethod
    def fileNameOnly(cls, path):
        return os.path.basename(path)
    
    @classmethod
    def dirname(cls,path):
        return os.path.dirname(path)
    
    @classmethod
    def ensureDirectory(cls, path, tail=None):
        # make sure the target dir is present
        if tail is not None:
            path = os.path.join(path, tail)
        if not os.access(path, os.F_OK):
            try:
                os.makedirs(path)
                os.chmod(path, 0o777) 
            except OSError as osError:
                logging.log(logging.ERROR, "target not created:" + path)
                logging.log(logging.ERROR, "Error: " + str(osError.strerror))
    
    @classmethod
    def __pairwise(cls,iterable):
        a, b = tee(iterable)
        next(b, None)
        return list(zip(a, b))    
    #logging rotation & compression
    @classmethod
    def compressor(cls,source, dest):
        with open(source,'rb') as srcFile:
            data=srcFile.read()
            bindata = bytearray(data)
            with gzip.open(dest,'wb') as gz:
                gz.write(bindata)
        os.remove(source)
    
    @classmethod
    def namer(self,name):
        return name+".gz"

    @classmethod
    def setupRotatingLogger(cls,logName,logConsole):
        '''
        Note: desktop file are opened by current user - active directory can not be used (Permissions) 
        '''
        logSize=5*1024*1024 #5MB
        if logConsole: #aka debug/development
            folder = OSTools.getActiveDirectory()    
        else:
            folder= OSTools.joinPathes(OSTools().getHomeDirectory(),".config",logName)
            OSTools.ensureDirectory(folder)
        logPath = OSTools.joinPathes(folder,logName+".log") 
        fh= RotatingFileHandler(logPath,maxBytes=logSize,backupCount=5)
        fh.rotator=OSTools.compressor
        fh.namer=OSTools.namer
        logHandlers=[]
        logHandlers.append(fh)
        if logConsole:
            logHandlers.append(logging.StreamHandler(sys.stdout))    
        logging.basicConfig(
            handlers=logHandlers,
            #level=logging.INFO
            level=logging.DEBUG,
            format='%(asctime)s %(levelname)s : %(message)s'
        )


    @classmethod
    def setLogLevel(cls,levelString):
        if levelString == "Debug":
            Log.setLevel(logging.DEBUG)
        elif levelString == "Info":
            Log.setLevel(logging.INFO)
        elif levelString == "Warning":
            Log.setLevel(logging.WARNING)
        elif levelString == "Error":
            Log.setLevel(logging.ERROR)

    #will not work in windows
    @classmethod
    def checkIfInstanceRunning(cls,moduleName):
        process = Popen(["ps aux |grep -v grep | grep "+moduleName],shell=True,stdout=subprocess.PIPE)
        result = process.communicate()
        rows = result[0].decode('utf8').splitlines()
        instanceCount =0
        for line in rows:
            if line:
                instanceCount+=1
        return instanceCount > 1


Log=logging.getLogger("OMOC")


class WorksheetReader():
    def __init__(self,filePath):
        self.file=filePath
    
    def importXls(self,adaptercls):
        filterFields = adaptercls.FIELDS
        wb = load_workbook(filename=self.file, read_only=True)
        sheets= wb.sheetnames
        ws=wb[sheets[0]]
        fullSheet=[]
        for row in ws.rows:
            singleRow=[]
            col=0
            for cell in row:
                if col in filterFields:
                    saveVal=cell.value if cell.value else ''
                    singleRow.append(saveVal)
                col +=1     
            obj=adaptercls()
            obj.row=singleRow
            if not obj.isBlocked():
                fullSheet.append(obj)        
            '''
            for cell in row:
                saveVal=cell.value if cell.value else '' 
                singleRow.append(saveVal)
            fullSheet.append(singleRow)
            '''
        cntMbrs=len(fullSheet)
        head=fullSheet.pop(0)
        cntCols = len(head.FIELDS)
        Log.info("Read %d entries, with %d columns each",cntMbrs,cntCols)
        wb.close()        
        return fullSheet   


class OmocAdapter():
    FIELDS = (1,2,3,5,6) #A=0
    ID=5
    BLOCKED=("Schulsport","Post-SV","gesperrt", "SVL")
    def __init(self,tupleRow):
        self.row=tupleRow()
        
    def fromDate(self):
        raw=self._cleanDate() #has space in front..
        hm = self.row[1]
        return datetime.strptime(hm+raw,"%H:%M %d.%m.%Y")

    def fromRoundDate(self):
        raw=self.fromDate()
        #lkr sometimes has 35. Round that to 30
        return raw - timedelta(minutes=raw.minute %10)
        
    def toDate(self):    
        raw=self._cleanDate()
        hm = self.row[2]
        return datetime.strptime(hm+raw,"%H:%M %d.%m.%Y")        
    
    def toRoundDate(self):
        raw=self.toDate()
        #lkr sometimes has 35. Round that to 30
        return raw - timedelta(minutes=raw.minute %10)     
    
    def timeDelta(self):
        return self.toRoundDate()-self.fromRoundDate()
    
    def _cleanDate(self):
        return self.row[0].split(',')[1]
    
    def isBlocked(self):
        key = self.row[3]
        for item in self.BLOCKED:
            if item in key:
                #print("Blocked:",key)
                return True
        return False
        
    def location(self):
        return self.row[4]
    
    def __str__(self):
        return "Omoc: %s [%s] - %s"%(self.fromDate(),self.fromRoundDate(),self.toDate())
    
    def asDisplayString(self):
        return "%s > %s"%(self.fromDate(),self.toDate())

class LkrAdapter():
    #03.07.2023-16:30-18:00-Berufsschule Weilheim Sporthalle-3-15
    FIELDS = (9,10,12,13,14,15) #A=0

    def __init(self,tupleRow):
        self.row=tupleRow()
        
    def fromDate(self):
        raw=self._cleanDate() #has space in front..
        hm = self.row[1]
        return datetime.strptime(hm+raw,"%H:%M %d.%m.%Y")
    
    def fromRoundDate(self):
        raw=self.fromDate()
        #lkr sometimes has 35. Round that to 30
        return raw - timedelta(minutes=raw.minute %10)
        
    def toDate(self):    
        raw=self._cleanDate()
        hm = self.row[2]
        return datetime.strptime(hm+raw,"%H:%M %d.%m.%Y")        
    
    def toRoundDate(self):
        raw=self.toDate()
        #lkr sometimes has 35. Round that to 30
        return raw - timedelta(minutes=raw.minute %10)    
    
    def timeDelta(self):
        return self.toRoundDate()-self.fromRoundDate()
    
    def _cleanDate(self):
        return " "+self.row[0]
    
    def location(self):
        return self.row[3]
    
    def timeUnits(self):
        return self.row[4]
    
    def price(self):
        return self.row[5]
    
    def isBlocked(self):
        return False #no rules
    
    def __str__(self):
        return "LKR: %s [%s] - %s "%(self.fromDate(),self.fromRoundDate(),self.toDate())
    
    def asDisplayString(self):
        return "%s > %s"%(self.fromDate(),self.toDate())
    

class BookingEntry():
    def __init__(self):
        self.date=None
        self.lkr=None
        self.omc=None
    
    def omcDisplay(self):
        if self.omc:
            return self.omc.asDisplayString()
        return "n.V"
    
    def lkrDisplay(self):
        if self.lkr:
            return self.lkr.asDisplayString()
        return "n.V"

    
class WorkSheetWriter():
    def __init__(self,filePath):
        self.file=filePath
    
    def export(self,title,header,crazyArray):
        #fontStyle = Font(size = "10")
        #ws['B3'].font = Font(bold=True)
        wb = Workbook()
        ws = wb.active
        ws.title=title
        ws.append([title])
        ws["A1"].font = Font(size = "14",bold=True)
        ws.append(header)        
        for idx in range(0,len(header)):
            char=chr(65+idx)
            ws[char+str(2)].font=Font(bold=True)

        for row in crazyArray:
            ws.append(row)
            
        self._setColumnDimensions(ws)

        wb.save(filename=self.file)
        
    def _setColumnDimensions(self,ws):
        dims={}
        for row in ws.rows:
            for cell in row:
                dims[cell.column_letter]=max((dims.get(cell.column_letter, 0), len(str(cell.value)))) 

        for col, value in dims.items():
            ws.column_dimensions[col].width = value                
    
class WorkSheetComparer():
    def __init__(self):
        self.omocSheet=None
        self.lkrSheet=None
        self.data=[] #Date, omoc,LKR
        self.statistics=() #count same,count different == Statistics 
        self.location = None #Kontext for the Writen, later. which location
        
    def readOmoc(self,omocPath):
        wr=WorksheetReader(omocPath)
        self.omocSheet = wr.importXls(OmocAdapter)
        self.omocSheet.sort(key=lambda adapter: adapter.fromDate() )
    
    def readLkr(self,lkrPath):
        wr=WorksheetReader(lkrPath)
        self.lkrSheet = wr.importXls(LkrAdapter)
        self.lkrSheet.sort(key=lambda adapter: adapter.fromDate() )        

    def compare(self):
        dic={} #Date ->bookingEntry
        for adapter in self.omocSheet:
            be = dic.get(adapter.fromRoundDate(),BookingEntry())
            be.omc=adapter
            dic[adapter.fromRoundDate()]=be
            
        for adapter in self.lkrSheet:
            be = dic.get(adapter.fromRoundDate(),BookingEntry())
            be.lkr=adapter
            dic[adapter.fromRoundDate()]=be

        same=0
        diff=0
        for key,bookingEntry in dic.items():
            if bookingEntry.omc is not None and bookingEntry.lkr is not None and bookingEntry.omc.timeDelta()==bookingEntry.lkr.timeDelta():
                same += 1
            else:
                #print(key,"O:",bookingEntry.omc," L:",bookingEntry.lkr)
                self.data.append([str(key),bookingEntry.omcDisplay(),bookingEntry.lkrDisplay()])
                diff += 1
     
        self.statistics=(same,diff)
        #print("Same: %d diff: %d"%(same,diff))
                
           
def test1():
    path = OSTools.getLocalPath(__file__)
    #p1 = OSTools.joinPathes(path, "testData", "OMOC Daten","20240116_FOS_OMOC.xlsx")
    #p2 = OSTools.joinPathes(path, "testData", "Landkreislisten","FOS-Landkreis.xlsx")
    #p1 = OSTools.joinPathes(path, "testData", "OMOC Daten","20240116_Jahnhalle_OMOC.xlsx")
    #p2 = OSTools.joinPathes(path, "testData", "Landkreislisten","Jahnhalle-Landkreis.xlsx")
    #p1 = OSTools.joinPathes(path, "testData", "OMOC Daten","20240116_Gymnasium_OMOC.xlsx")
    #p2 = OSTools.joinPathes(path, "testData", "Landkreislisten","GYMLandkreis.xlsx")
    p1 = OSTools.joinPathes(path, "testData", "OMOC Daten","20240116_Hallenbad_OMOC.xlsx")
    p2 = OSTools.joinPathes(path, "testData", "Landkreislisten","Schwimmbad-Landkreis.xlsx")
    
    wc = WorkSheetComparer()
    wc.readOmoc(p1)
    wc.readLkr(p2)
    wc.compare()
    #testest
    target=OSTools.joinPathes(path,"testData","Frigin.xlsx")
    wr=WorkSheetWriter(target)
    wr.export("Hallenbad",("Datum","OMOC","LKR"),wc.data)
 


if __name__ == '__main__':
    test1()